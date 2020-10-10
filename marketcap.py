import pandas as pd
import numpy as np

from openpyxl import load_workbook
wb = load_workbook('/Users/maria/Job/Python/Ws.xlsx')  # Work Book file that I have 
ws = wb['Sheet2']  # Work Sheet
column = ws['C']  # Column
tickers_list = [column[x].value for x in range(len(column))]

print(*tickers_list, sep='\n')  #επιβεβαιώνω ότι λειτουργεί σωστά και τυπώνει τα tickers από excel
print(len(tickers_list))  #επιβεβαιώνω ότι το μέγεθος των στοιχείων είναι σωστό

#extract the marketcaps
import pandas_datareader as web
tickers = tickers_list[1:]
print(tickers)
market_cap_data= web.get_quote_yahoo(tickers)['marketCap']
print(market_cap_data)

